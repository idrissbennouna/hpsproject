import io
import os
import shutil
import html
import re
import tempfile
import hashlib
import uuid
from pathlib import Path
from typing import List, Optional, Dict, Any

from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.concurrency import run_in_threadpool
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse
from pydantic import BaseModel, Field
from dotenv import load_dotenv

# Chargement des variables d'environnement
load_dotenv()

# Cache mémoire pour la documentation enrichie par LLM (clé : (function_name, session_id))
_FUNC_DOC_CACHE: Dict[tuple, dict] = {}

# Importations des applications LangGraph
from app.core.agent_graph import compliance_agent_app
from app.core.validation_agent_graph import validation_agent_app

import logging
import traceback

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("compliance_verifier")

app = FastAPI(
    title="ComplianceVerifier API - HPS",
    description="Backend d'analyse agentique de logs monétiques et de spécifications",
    version="1.0.0"
)

@app.exception_handler(Exception)
async def global_exception_handler(request, exc: Exception):
    tb_str = "".join(traceback.format_exception(type(exc), exc, exc.__traceback__))
    logger.error(f"[GLOBAL_EXCEPTION_HANDLER] Unhandled exception occurred at path: {request.url.path}\n{tb_str}")
    print(f"🔥 [CRITICAL GLOBAL UNHANDLED EXCEPTION] at {request.url.path}:\n{tb_str}")
    return JSONResponse(
        status_code=500,
        content={
            "detail": f"Internal Server Error: {str(exc)}",
            "traceback": tb_str
        }
    )

# 1. Configuration du CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://127.0.0.1:5173", "*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# R�solution s�curis�e du dossier de stockage
CURRENT_DIR = Path(__file__).resolve().parent
STORAGE_DIR = CURRENT_DIR / "storage"
STORAGE_DIR.mkdir(parents=True, exist_ok=True)


# --- Utilitaires Internes ---

def _normalize_response_text(content: any) -> str:
    """Normalise le contenu textuel renvoy� par les agents (liste, dict, objet ou None)."""
    if content is None:
        return ""
    if isinstance(content, list):
        parts = []
        for block in content:
            if isinstance(block, str):
                parts.append(block)
            elif isinstance(block, dict):
                parts.append(block.get("text", str(block)))
            elif hasattr(block, "text"):
                parts.append(getattr(block, "text", str(block)))
            else:
                parts.append(str(block))
        return "\n".join(parts)
    return str(content)


def _generate_reportlab_pdf(final_report_data: any, agent_assigned: str, filename: str, pdf_path: Path) -> None:
    """Génère le rapport PDF sécurisé avec la charte graphique HPS à partir d'un rapport structuré ou brut."""
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors

    doc = SimpleDocTemplate(
        str(pdf_path),
        pagesize=letter,
        rightMargin=54, leftMargin=54,
        topMargin=54, bottomMargin=54
    )
    story = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'DocTitle', parent=styles['Heading1'], fontSize=22, leading=26,
        textColor=colors.HexColor('#0F172A'), spaceAfter=5, alignment=0
    )
    section_style = ParagraphStyle(
        'SectionTitle', parent=styles['Heading2'], fontSize=13, leading=17,
        textColor=colors.HexColor('#1E3A8A'), spaceBefore=14, spaceAfter=8, keepWithNext=True
    )
    body_style = ParagraphStyle(
        'ReportBody', parent=styles['Normal'], fontSize=9, leading=13,
        textColor=colors.HexColor('#334155'), spaceAfter=6
    )
    bold_style = ParagraphStyle(
        'ReportBold', parent=body_style, fontName='Helvetica-Bold'
    )
    header_text_style = ParagraphStyle(
        'HeaderText', parent=styles['Normal'], fontSize=8, leading=10,
        textColor=colors.HexColor('#64748B'), alignment=2
    )
    table_header_style = ParagraphStyle(
        'TableHeader', parent=body_style, fontSize=9, textColor=colors.white, fontName='Helvetica-Bold'
    )

    logo_path = STORAGE_DIR / "HPS_logo.png"
    if logo_path.exists():
        logo_img = Image(str(logo_path), width=100, height=35)
        logo_img.hAlign = 'LEFT'
        story.append(logo_img)
        story.append(Spacer(1, 10))

    story.append(Paragraph("<b>HPS ComplianceAI Platform</b> — Automated Log Verification Terminal", header_text_style))
    story.append(Spacer(1, 10))
    story.append(Paragraph("Rapport de Validation Monétique", title_style))
    story.append(Paragraph("<font color='#64748B'><i>Analyse automatisée de conformité des traces d'autorisation</i></font>", body_style))
    story.append(Spacer(1, 10))

    meta_data = [
        [Paragraph("Paramètre d'Audit", table_header_style), Paragraph("Valeur / Référence", table_header_style)],
        [Paragraph("Fichier Trace Source", body_style), Paragraph(html.escape(filename), body_style)],
        [Paragraph("Référence Spécification", body_style), Paragraph("Spec_PowerCARD.xlsx", body_style)],
        [Paragraph("Moteur d'Analyse", body_style), Paragraph(html.escape(agent_assigned), body_style)],
        [Paragraph("Statut Système", body_style), Paragraph("Analyse Terminée (200 OK)", body_style)]
    ]

    meta_table = Table(meta_data, colWidths=[160, 340])
    meta_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (1, 0), colors.HexColor('#1E3A8A')),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('LINEBELOW', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
    ]))

    story.append(meta_table)
    story.append(Spacer(1, 12))

    # Support raw fallback or formatted dict
    if isinstance(final_report_data, dict) and "summary" in final_report_data:
        summary = final_report_data.get("summary", {})
        story.append(Paragraph("Synthèse de l'Analyse", section_style))
        summary_data = [
            [Paragraph("Total Transactions", table_header_style), Paragraph("Alertes / Suspectes", table_header_style), Paragraph("Approuvées", table_header_style), Paragraph("Déclinées", table_header_style)],
            [
                Paragraph(str(summary.get("total_transactions", 0)), body_style),
                Paragraph(f"<font color='#DC2626'><b>{summary.get('suspicious_count', 0)}</b></font>", body_style),
                Paragraph(f"<font color='#16A34A'><b>{summary.get('approved_count', 0)}</b></font>", body_style),
                Paragraph(f"<font color='#DC2626'><b>{summary.get('declined_count', 0)}</b></font>", body_style)
            ]
        ]
        sum_table = Table(summary_data, colWidths=[125, 125, 125, 125])
        sum_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A8A')),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ]))
        story.append(sum_table)
        story.append(Spacer(1, 10))

        # Transactions
        txs = final_report_data.get("transactions", [])
        if txs:
            story.append(Paragraph("Transactions Analysées", section_style))
            for tx in txs:
                tx_id = html.escape(str(tx.get("transaction_id", "")))
                status = "Approuvée" if tx.get("approval_status") == "approved" else "Déclinée"
                status_color = "#16A34A" if tx.get("approval_status") == "approved" else "#DC2626"
                susp_label = " [⚠️ ALERTE DETECTEE]" if tx.get("is_suspicious") else ""
                
                header_p = Paragraph(f"<b>{tx_id}</b> - Statut: <font color='{status_color}'><b>{status}</b></font>{susp_label}", body_style)
                story.append(header_p)
                
                details = f"PAN: {html.escape(str(tx.get('pan_masked', '')))} | STAN: {html.escape(str(tx.get('stan', '')))} | RRN: {html.escape(str(tx.get('rrn', '')))} | Code Rep: {html.escape(str(tx.get('response_code', '')))}"
                story.append(Paragraph(f"<font color='#64748B'>{details}</font>", body_style))
                
                alerts = tx.get("alerts", [])
                if alerts:
                    story.append(Paragraph(f"<b>Alertes:</b> <font color='#DC2626'>{html.escape(', '.join(alerts))}</font>", body_style))
                
                chronology = tx.get("chronology", [])
                if chronology:
                    for step in chronology:
                        story.append(Paragraph(f"&bull; {html.escape(str(step))}", body_style))
                story.append(Spacer(1, 6))

        # Field Analysis — NON-CONFORMITES UNIQUEMENT
        fa = final_report_data.get("field_analysis", [])
        no_violations = final_report_data.get("no_field_violations", False)
        story.append(Paragraph("Analyse de Conformité des Champs", section_style))
        if not fa or no_violations:
            story.append(Paragraph(
                "✅ Aucune non-conformité de format détectée pour cette analyse.",
                ParagraphStyle('GreenNote', parent=body_style, textColor=colors.HexColor('#16A34A'))
            ))
        else:
            violation_data = [
                [
                    Paragraph("Champ ISO", table_header_style),
                    Paragraph("Type Attendu (source)", table_header_style),
                    Paragraph("Valeur Observée", table_header_style),
                    Paragraph("Nature de la Non-Conformité", table_header_style),
                ]
            ]
            for field in fa:
                f_num = html.escape(str(field.get("field_number", "")))
                f_name = html.escape(str(field.get("field_name", "")))
                f_type = html.escape(str(field.get("expected_type", "")))
                f_src = html.escape(str(field.get("source", "")))
                f_val = html.escape(str(field.get("observed_value", field.get("value", ""))))
                f_nc = html.escape(str(field.get("non_conformity_type", "")))
                violation_data.append([
                    Paragraph(f"<b>{f_num}</b>\n{f_name}", body_style),
                    Paragraph(f"{f_type}\n<font color='#64748B'>{f_src}</font>", body_style),
                    Paragraph(f"<code>{f_val}</code>", body_style),
                    Paragraph(f"<font color='#DC2626'>{f_nc}</font>", body_style),
                ])
            viol_table = Table(violation_data, colWidths=[90, 120, 120, 170])
            viol_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A8A')),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#FEF2F2')),
            ]))
            story.append(viol_table)
        story.append(Spacer(1, 10))

    else:
        raw_text = final_report_data.get("raw_fallback", "") if isinstance(final_report_data, dict) else str(final_report_data)
        story.append(Paragraph("Résultats de l'Analyse Agentique", section_style))
        for line in raw_text.split("\n"):
            if line.strip():
                safe_line = html.escape(line)
                if safe_line.strip().startswith("#### "):
                    safe_line = "<b>" + safe_line.replace("#### ", "", 1) + "</b>"
                if safe_line.strip().startswith("* "):
                    safe_line = safe_line.replace("* ", "&bull; ", 1)
                safe_line = re.sub(r'\*\*(.*?)\*\*', r'<b>\1</b>', safe_line)
                safe_line = re.sub(r'`(.*?)`', r'<b>\1</b>', safe_line)
                story.append(Paragraph(safe_line, body_style))

    doc.build(story)


def _process_and_index_pdf(file_bytes: bytes, filename: str, session_id: str) -> dict:
    """
    Traite et indexe un fichier PDF pour la session de validation.
    
    LOGIQUE DE DÉDUPLICATION PAR HASH (SHA-256) ET CHUNKING PAR FENÊTRE GLISSANTE :
    1. Calcule le hash SHA-256 du fichier PDF. Si ce hash existe déjà dans pgvector,
       réutilise tous ses chunks existants et met à jour session_id / last_accessed
       sans ré-appeler l'API Gemini (économie de quota).
    2. Pour la découpe (chunking) : afin d'éviter la perte de contexte entre deux pages
       consécutives (ex: commande 'EC' sur la page N et table des codes d'erreur 'ED'
       sur la page N+1), fusionne le texte des pages contiguës dans une fenêtre glissante
       (Page N + Page N+1) AVANT d'appliquer le RecursiveCharacterTextSplitter.
    3. Extrait par regex les codes de commande / réponse (ex: 'EC (ED)') pour alimenter les
       métadonnées 'command_code' et 'response_code' sur chaque chunk final.
    """
    import pdfplumber
    from datetime import datetime, timezone
    from langchain_core.documents import Document
    from langchain_text_splitters import RecursiveCharacterTextSplitter
    from app.rag.retriever import (
        get_session_vectorstore,
        delete_session_documents,
        batch_add_documents,
        QuotaExhaustedError,
        find_chunks_by_file_hash,
    )
    from app.core.config import (
        MAX_PDF_PAGES, 
        PDF_CHUNK_SIZE, 
        PDF_CHUNK_OVERLAP, 
        EMBEDDING_BATCH_SIZE, 
        EMBEDDING_BATCH_DELAY_SECONDS,
        CHUNKING_VERSION
    )

    raw_pages = []
    total_pages = 0
    file_hash = hashlib.sha256(file_bytes).hexdigest()
    existing_info = find_chunks_by_file_hash(file_hash)
    existing_chunk_indices = existing_info.get("existing_chunk_indices", set())

    try:
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            total_pages = len(pdf.pages)

            if total_pages > MAX_PDF_PAGES:
                return {
                    "success": False,
                    "error_type": "too_many_pages",
                    "filename": filename,
                    "message": f"Le nombre de pages ({total_pages}) dépasse la limite autorisée de {MAX_PDF_PAGES} pages."
                }

            delete_session_documents(session_id)

            for idx, page in enumerate(pdf.pages, 1):
                page_text = page.extract_text() or ""
                if page_text.strip():
                    raw_pages.append((idx, page_text))
    except Exception as pdf_err:
        return {
            "success": False,
            "error_type": "pdf_error",
            "filename": filename,
            "message": f"Fichier PDF invalide ou illisible : {pdf_err}",
            "total_pages": 0,
            "extracted_text": ""
        }

    pages_with_text = len(raw_pages)
    if total_pages > 0 and (pages_with_text == 0 or (pages_with_text / total_pages) < 0.5):
        msg = (
            "Aucun texte extractible trouvé dans le document PDF. Le fichier est probablement scanné."
            if pages_with_text == 0 else
            f"Seulement {pages_with_text}/{total_pages} pages contiennent du texte extractible. Le document semble majoritairement scanné."
        )
        return {
            "success": False,
            "error_type": "scanned_pdf",
            "filename": filename,
            "message": msg,
            "total_pages": total_pages,
            "extracted_text": ""
        }

    # 1. Fusion par fenêtre glissante optimisée (Page N + pont de contexte de Page N+1)
    documents = []
    now_iso = datetime.now(timezone.utc).isoformat()
    if len(raw_pages) == 1:
        p_num, p_text = raw_pages[0]
        documents.append(Document(
            page_content=f"--- Page {p_num} ---\n{p_text}",
            metadata={
                "session_id": session_id,
                "file_hash": file_hash,
                "source": filename,
                "page": str(p_num),
                "created_at": now_iso,
                "last_accessed": now_iso
            }
        ))
    else:
        for i in range(len(raw_pages)):
            p1_num, p1_text = raw_pages[i]
            if i + 1 < len(raw_pages):
                p2_num, p2_text = raw_pages[i + 1]
                bridge = p2_text[:PDF_CHUNK_OVERLAP]
                combined_content = f"--- Page {p1_num} ---\n{p1_text}\n\n--- Début page suivante (contexte) ---\n{bridge}"
                page_meta = f"{p1_num}-{p2_num}"
            else:
                combined_content = f"--- Page {p1_num} ---\n{p1_text}"
                page_meta = str(p1_num)

            documents.append(Document(
                page_content=combined_content,
                metadata={
                    "session_id": session_id,
                    "file_hash": file_hash,
                    "source": filename,
                    "page": page_meta,
                    "created_at": now_iso,
                    "last_accessed": now_iso
                }
            ))

    text_splitter = RecursiveCharacterTextSplitter(chunk_size=PDF_CHUNK_SIZE, chunk_overlap=PDF_CHUNK_OVERLAP)
    chunks = text_splitter.split_documents(documents)

    # Regex d'extraction des paires commande/réponse (ex: "EC (ED)" ou "EC/ED")
    cmd_pair_regex = re.compile(r"\b([A-Z0-9]{2})\s*(?:\/|\s+and\s+|\s*\(\s*)\s*([A-Z0-9]{2})\s*\)?", re.IGNORECASE)
    # Regex d'extraction spécifique des Response Codes (ex: "Response Code 2 A Value 'ED'")
    resp_code_val_regex = re.compile(r"Response Code\s+2\s*A\s+Value\s*'(\w{2})'", re.IGNORECASE)

    for c_idx, chunk in enumerate(chunks):
        chunk.metadata["chunk_index"] = c_idx
        chunk.metadata["chunking_version"] = CHUNKING_VERSION
        chunk.metadata["created_at"] = now_iso
        chunk.metadata["last_accessed"] = now_iso

        # Extraction regex des codes commande / réponse pour metadata pgvector
        match_cmd = cmd_pair_regex.search(chunk.page_content)
        if match_cmd:
            chunk.metadata["command_code"] = match_cmd.group(1).upper()
            chunk.metadata["response_code"] = match_cmd.group(2).upper()

        match_resp = resp_code_val_regex.search(chunk.page_content)
        if match_resp:
            chunk.metadata["response_code"] = match_resp.group(1).upper()

        # Signal supplémentaire: contains_error_table
        content_lower = chunk.page_content.lower()
        error_table_keywords = ["parity error", "pin block format has been disabled", "response code", "error code"]
        chunk.metadata["contains_error_table"] = any(kw in content_lower for kw in error_table_keywords)

    if existing_info.get("found") and len(existing_chunk_indices) >= len(chunks):
        from app.rag.retriever import update_session_id_for_file_hash
        update_session_id_for_file_hash(file_hash, session_id)
        print(f"[REUSE_HASH] File '{filename}' (hash: {file_hash[:10]}...) completely embedded ({len(existing_chunk_indices)}/{len(chunks)} chunks). Reusing for session '{session_id}'.")
        return {
            "success": True,
            "reused": True,
            "file_hash": file_hash,
            "filename": filename,
            "message": "Fichier éphémère réutilisé sans ré-embedding.",
            "total_pages": total_pages,
            "extracted_text": f"[Document PDF indexé dans la base vectorielle RAG de session. Total : {total_pages} pages.]"
        }

    if existing_info.get("version_mismatch"):
        from app.rag.retriever import purge_chunks_by_file_hash
        purge_chunks_by_file_hash(file_hash)
        existing_chunk_indices = set()

    missing_chunks = [c for c in chunks if c.metadata["chunk_index"] not in existing_chunk_indices]

    if existing_chunk_indices:
        print(f"[RESUME_HASH] File '{filename}' (hash: {file_hash[:10]}...): Resuming ingestion. {len(existing_chunk_indices)}/{len(chunks)} chunks already present, embedding remaining {len(missing_chunks)} chunks.")

    session_db = get_session_vectorstore()
    try:
        indexed_count = batch_add_documents(
            session_db,
            missing_chunks,
            batch_size=EMBEDDING_BATCH_SIZE,
            inter_batch_delay=EMBEDDING_BATCH_DELAY_SECONDS,
            session_id=session_id
        )
        print(f"[INDEX_SUCCESS] Session '{session_id}': {indexed_count}/{len(missing_chunks)} missing chunks indexed successfully.")
    except QuotaExhaustedError as qe:
        return {
            "success": False,
            "error_type": "quota_exhausted",
            "filename": filename,
            "file_hash": file_hash,
            "indexed": qe.indexed_count,
            "total": len(missing_chunks),
            "message": f"Quota API atteint après l'indexation de {qe.indexed_count}/{len(missing_chunks)} chunks."
        }
    except Exception as e:
        return {
            "success": False,
            "error_type": "ingestion_failed",
            "filename": filename,
            "file_hash": file_hash,
            "message": f"Erreur lors de l'indexation vectorielle du PDF : {str(e)}"
        }

    return {
        "success": True,
        "reused": False,
        "file_hash": file_hash,
        "filename": filename,
        "message": "Fichier éphémère extrait et chargé dans la session avec succès.",
        "total_pages": total_pages,
        "extracted_text": f"[Document PDF indexé dans la base vectorielle RAG de session. Total : {total_pages} pages.]"
    }


@app.get("/")
def read_root():
    return {"status": "online", "message": "Serveur ComplianceVerifier oprationnel"}


@app.post("/api/v1/logs/analyze")
async def analyze_logs(
    user_prompt: str = Form(...),
    file: UploadFile = File(...),
    spec_file: Optional[UploadFile] = File(None, alias="doc_file"),
    doc_file: Optional[UploadFile] = File(None),
    spec_file_hash: Optional[str] = Form(None),
    job_id: Optional[str] = Form(None)
):
    """
    Reçoit un fichier de traces (.TXT, .LOG, .TRC, .DAT), analyse son contenu
    via le graphe d'agents et produit un rapport PDF.
    Accepte facultativement un fichier de spécification (PDF) à uploader OU
    un file_hash d'un document déjà indexé (réutilisation sans ré-embedding).
    """
    safe_filename = Path(file.filename or "upload").name
    filename_lower = safe_filename.lower()
    is_trace_file = any(filename_lower.endswith(ext) for ext in ('.txt', '.log', '.trc', '.dat')) or '.trc' in filename_lower

    if not is_trace_file:
        raise HTTPException(status_code=400, detail="Seuls les fichiers de traces (.TXT, .LOG, .TRC, .DAT) sont acceptés.")

    active_job_id = job_id or f"job_{uuid.uuid4().hex[:12]}"
    from app.services.job_tracker import create_job, update_job
    create_job(active_job_id, job_type="log_analysis")
    update_job(active_job_id, stage="parsing_trace", detail="Réception et préparation du fichier de traces...", progress_pct=5)

    target_file_path = STORAGE_DIR / safe_filename

    try:
        with target_file_path.open("wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
    except Exception as e:
        update_job(active_job_id, stage="error", detail=f"Échec d'écriture : {str(e)}", error=str(e))
        raise HTTPException(status_code=500, detail=f"Échec de l'écriture du fichier : {str(e)}")

    effective_spec = spec_file or doc_file
    doc_session_id = active_job_id

    # --- Réutilisation d'un document déjà indexé via son file_hash ---
    if spec_file_hash and spec_file_hash.strip() and not (effective_spec and effective_spec.filename):
        try:
            from app.rag.retriever import update_session_id_for_file_hash
            update_session_id_for_file_hash(spec_file_hash.strip(), doc_session_id)
            print(f"[REUSE_SPEC] Spec document (hash: {spec_file_hash[:10]}...) réutilisé pour la session '{doc_session_id}' sans ré-embedding.")
            update_job(active_job_id, stage="embedding_pdf", detail="Document de spécification existant associé à la session.", progress_pct=20)
        except Exception as hash_err:
            print(f"[WARN] Impossible de réutiliser le document via hash '{spec_file_hash[:10]}': {hash_err}")

    elif effective_spec and effective_spec.filename:
        try:
            update_job(active_job_id, stage="embedding_pdf", detail="Extraction et indexation du document de spécification...", progress_pct=15)
            spec_bytes = await effective_spec.read()
            if spec_bytes:
                safe_spec_name = Path(effective_spec.filename).name
                pdf_res = await run_in_threadpool(_process_and_index_pdf, spec_bytes, safe_spec_name, doc_session_id)
                if not pdf_res.get("success"):
                    print(f"[WARN] Indexation du document de spec échouée : {pdf_res.get('message')}")
                    logger.warning(f"Indexation document spec échouée pour {active_job_id}: {pdf_res.get('message')}")
        except Exception as spec_err:
            print(f"[WARN] Erreur lors de la lecture/indexation du document de spec : {str(spec_err)}")
            logger.warning(f"Erreur lecture doc spec pour {active_job_id}: {str(spec_err)}")

    try:
        graph_inputs = {
            "user_prompt": user_prompt,
            "file_name": safe_filename,
            "doc_session_id": doc_session_id,
            "current_agent": "",
            "rag_context": "",
            "log_data_json": "",
            "final_response": ""
        }

        update_job(active_job_id, stage="generating_report", detail="Exécution du graphe multi-agents (LogStory & Compliance)...", progress_pct=50)
        execution_result = await run_in_threadpool(compliance_agent_app.invoke, graph_inputs)

        raw_report = execution_result.get("final_response", {"raw_fallback": "Aucun rapport généré."})
        agent_assigned = execution_result.get("current_agent", "ComplianceAuditorAgent")

        # PART 1: Unique PDF filename per request to prevent stale file caching
        unique_suffix = uuid.uuid4().hex[:8]
        pdf_filename = f"Rapport_{Path(safe_filename).stem}_{unique_suffix}.pdf"
        pdf_path = STORAGE_DIR / pdf_filename

        update_job(active_job_id, stage="building_pdf", detail="Génération du rapport PDF...", progress_pct=85)

        pdf_generation_failed = False
        pdf_error_detail = None

        try:
            await run_in_threadpool(_generate_reportlab_pdf, raw_report, agent_assigned, safe_filename, pdf_path)
            print(f"[OK] PDF principal généré avec succès : {pdf_filename}")
        except Exception as pdf_err:
            tb_err = traceback.format_exc()
            print(f"[ERROR] Erreur lors de la génération du PDF principal ({pdf_filename}) : {pdf_err}\n{tb_err}")
            logger.error(f"Échec génération PDF principal pour {pdf_filename}: {pdf_err}\n{tb_err}")
            
            try:
                print(f"[WARN] Tentative de génération du PDF fallback Mode Restauré pour {pdf_filename}...")
                from reportlab.lib.pagesizes import letter
                from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
                from reportlab.lib.styles import getSampleStyleSheet

                doc_fail = SimpleDocTemplate(str(pdf_path), pagesize=letter)
                styles_fail = getSampleStyleSheet()
                story_fail = [
                    Paragraph("<b>Rapport d'Audit (Mode Restauré)</b>", styles_fail['Heading1']),
                    Spacer(1, 15)
                ]
                fallback_str = raw_report.get("raw_fallback", str(raw_report)) if isinstance(raw_report, dict) else str(raw_report)
                for raw_line in fallback_str.split("\n"):
                    if raw_line.strip():
                        story_fail.append(Paragraph(html.escape(raw_line), styles_fail['Normal']))
                doc_fail.build(story_fail)
                print(f"[OK] PDF fallback Mode Restauré généré avec succès : {pdf_filename}")
            except Exception as final_err:
                pdf_generation_failed = True
                pdf_error_detail = str(final_err)
                tb_final = traceback.format_exc()
                print(f"[CRITICAL] Échec critique du fallback PDF ({pdf_filename}) : {final_err}\n{tb_final}")
                logger.critical(f"Échec critique du fallback PDF ({pdf_filename}): {final_err}\n{tb_final}")

        response_data = {
            "success": True,
            "job_id": active_job_id,
            "agent_assigned": agent_assigned,
            "log_chronology": execution_result.get("log_data_json"),
            "report": raw_report,
            "analysis_report": raw_report,
            "pdf_filename": pdf_filename,
            "pdf_generation_failed": pdf_generation_failed,
            "pdf_error_detail": pdf_error_detail
        }

        update_job(active_job_id, stage="done", detail="Analyse et génération terminées avec succès !", progress_pct=100, result=response_data)

        # PARTIE A — Sauvegarde automatique de la conversation "logs"
        try:
            from app.services.conversation_history import create_conversation
            _conv_title = Path(safe_filename).stem[:60]
            create_conversation(
                agent_type="logs",
                title=_conv_title,
                messages=[{"role": "user", "content": user_prompt}],
                result=response_data,
                conv_id=active_job_id,
            )
        except Exception as _conv_err:
            print(f"[WARN] Impossible de sauvegarder la conversation logs : {_conv_err}")

        return response_data

    except Exception as e:
        update_job(active_job_id, stage="error", detail=f"Erreur agentique : {str(e)}", error=str(e))
        from app.services.llm_util import GeminiOverloadedError, GeminiQuotaExhaustedError
        real_exc = e.__cause__ or e
        if isinstance(real_exc, GeminiOverloadedError):
            raise HTTPException(status_code=503, detail="Le service Gemini est temporairement surchargé. Veuillez réessayer dans quelques instants.")
        elif isinstance(real_exc, GeminiQuotaExhaustedError):
            raise HTTPException(status_code=429, detail="Le quota de l'API Gemini a été atteint. Veuillez réessayer plus tard.")
        raise HTTPException(status_code=500, detail=f"Erreur lors de l'analyse agentique : {str(e)}")


@app.get("/api/v1/jobs/{job_id}/status")
def get_job_status(job_id: str):
    """Retourne le statut d'avancement d'un traitement en cours."""
    from app.services.job_tracker import get_job
    job = get_job(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job non trouvé.")
    return job


@app.get("/api/v1/logs/download-pdf")
async def download_pdf(filename: Optional[str] = None):
    """Permet le téléchargement du rapport PDF généré."""
    if not filename:
        raise HTTPException(status_code=400, detail="Le nom du fichier PDF doit être spécifié via le paramètre 'filename'.")
    
    safe_filename = Path(filename).name
    pdf_path = STORAGE_DIR / safe_filename

    if pdf_path.exists() and pdf_path.stat().st_size > 100:
        return FileResponse(
            path=str(pdf_path),
            filename=safe_filename,
            media_type="application/pdf"
        )

    raise HTTPException(
        status_code=404,
        detail="Aucun rapport d'analyse disponible sous ce nom. Veuillez ex�cuter l'analyse au pr�alable."
    )


class ValidationRequest(BaseModel):
    question: str
    chat_history: Optional[List[dict]] = Field(default_factory=list)
    session_id: Optional[str] = None
    conv_id: Optional[str] = None  # PARTIE A : identifiant de conversation docs


@app.post("/api/v1/validation/ask")
async def ask_validation_agent(request: ValidationRequest):
    """Interroge l'agent de validation sur les sp�cifications PowerCARD."""
    question = request.question.strip()
    if not question:
        raise HTTPException(status_code=400, detail="La question ne peut pas �tre vide.")

    try:
        inputs = {
            "user_question": question,
            "chat_history": request.chat_history,
            "rag_context": "",
            "final_response": "",
            "session_id": request.session_id or "",
            "sources": []
        }

        result = await run_in_threadpool(validation_agent_app.invoke, inputs)
        final_response = _normalize_response_text(result.get("final_response", ""))

        raw_sources = result.get("sources", [])
        sources = [s["label"] for s in raw_sources if isinstance(s, dict) and "label" in s] if raw_sources else []

        # PARTIE A — Sauvegarde automatique de la conversation "docs"
        saved_conv_id = None
        try:
            from app.services.conversation_history import create_conversation, update_conversation, get_conversation
            _question_preview = question[:50] + ("..." if len(question) > 50 else "")
            _all_messages = list(request.chat_history or []) + [
                {"role": "user", "content": question},
                {"role": "assistant", "content": final_response, "sources": sources},
            ]
            _cid = request.conv_id
            if _cid and get_conversation(_cid):
                update_conversation(_cid, messages=_all_messages)
                saved_conv_id = _cid
            else:
                _new_conv = create_conversation(
                    agent_type="docs",
                    title=_question_preview,
                    messages=_all_messages,
                )
                saved_conv_id = _new_conv["id"] if _new_conv else None
        except Exception as _conv_err:
            print(f"[WARN] Impossible de sauvegarder la conversation docs : {_conv_err}")

        return {
            "response": final_response,
            "sources": sources,
            "conv_id": saved_conv_id,
        }

    except Exception as e:
        from app.services.llm_util import GeminiOverloadedError, GeminiQuotaExhaustedError
        real_exc = e.__cause__ or e
        if isinstance(real_exc, GeminiOverloadedError):
            raise HTTPException(status_code=503, detail="Le service Gemini est temporairement surcharg�. Veuillez r�essayer dans quelques instants.")
        elif isinstance(real_exc, GeminiQuotaExhaustedError):
            raise HTTPException(status_code=429, detail="Le quota de l'API Gemini a �t� atteint. Veuillez r�essayer plus tard.")
        raise HTTPException(
            status_code=500,
            detail=f"Erreur lors de l'ex�cution de l'agent de validation : {str(e)}"
        )


@app.post("/api/v1/validation/upload")
async def upload_validation_file(
    file: UploadFile = File(...),
    session_id: str = Form(...)
):
    """
    Re�oit un fichier (.TXT, .PDF, .XLSX), extrait son texte brut,
    et l'enregistre en m�moire sous le session_id fourni.
    """
    safe_filename = Path(file.filename or "upload").name
    filename_lower = safe_filename.lower()

    is_trace_file = any(filename_lower.endswith(ext) for ext in ('.txt', '.log', '.trc', '.dat')) or '.trc' in filename_lower
    is_pdf_file = filename_lower.endswith('.pdf')
    is_excel_file = filename_lower.endswith('.xlsx') or filename_lower.endswith('.xls')

    if not (is_trace_file or is_pdf_file or is_excel_file):
        raise HTTPException(
            status_code=400,
            detail="Seuls les formats de fichiers de trace (.TXT, .LOG, .TRC), .PDF et .XLSX sont support�s."
        )

    try:
        from app.services.session_storage import add_session_file, compute_file_stats
        from app.core.config import MAX_UPLOAD_SIZE_BYTES, MAX_UPLOAD_SIZE_MB

        file_bytes = await file.read()

        if len(file_bytes) > MAX_UPLOAD_SIZE_BYTES:
            raise HTTPException(
                status_code=400,
                detail=f"La taille du fichier ({len(file_bytes) / (1024*1024):.1f} Mo) d�passe la limite autoris�e de {MAX_UPLOAD_SIZE_MB} Mo."
            )

        content = ""
        parsed_log_successfully = False
        file_hash = None

        if is_trace_file:
            with tempfile.NamedTemporaryFile(delete=False, suffix=".TXT") as temp_file:
                temp_file.write(file_bytes)
                temp_path = temp_file.name

            try:
                from app.services.log_parser import parse_and_format_log_file
                content = parse_and_format_log_file(temp_path, mode="compact")
                if content:
                    parsed_log_successfully = True
                else:
                    content = file_bytes.decode("utf-8", errors="ignore")
                    parsed_log_successfully = False
            finally:
                if os.path.exists(temp_path):
                    os.remove(temp_path)

        elif is_pdf_file:
            pdf_res = await run_in_threadpool(_process_and_index_pdf, file_bytes, safe_filename, session_id)
            if not pdf_res["success"]:
                status_code = 503 if pdf_res.get("error_type") == "quota_exhausted" else 400
                return JSONResponse(status_code=status_code, content=pdf_res)
            content = pdf_res["extracted_text"]
            file_hash = pdf_res.get("file_hash")

        elif is_excel_file:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            sheet_blocks = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_blocks.append(f"--- Onglet Excel : {sheet_name} ---")
                for r_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                    row_str = " | ".join([str(val).strip() for val in row if val is not None])
                    if row_str.strip():
                        sheet_blocks.append(f"Ligne {r_idx}: {row_str}")
            content = "\n".join(sheet_blocks)

        full_stats = compute_file_stats(content)
        is_rag_file = is_pdf_file

        if is_rag_file or parsed_log_successfully:
            full_stats["truncated_for_llm"] = False
        else:
            MAX_CHARACTERS = 15000
            total_raw_lines = len(content.splitlines())
            was_truncated = len(content) > MAX_CHARACTERS
            full_stats["truncated_for_llm"] = was_truncated
            full_stats["total_raw_lines"] = total_raw_lines
            full_stats["max_character_limit"] = MAX_CHARACTERS
            if was_truncated:
                content = content[:MAX_CHARACTERS] + (
                    f"\n\n[AVERTISSEMENT SYSTEME DE TRONCATURAGE TECHNIQUE : Le parsing structuré par session/transaction a échoué (aucune session reconnue). "
                    f"Le fichier brut de trace contient {total_raw_lines} lignes et {full_stats['char_count']} caractères. "
                    f"Afin d'éviter le dépassement de la fenêtre de contexte LLM, le texte brut non structuré a été tronqué aux {MAX_CHARACTERS} premiers caractères.]"
                )

        add_session_file(session_id, safe_filename, content, full_stats=full_stats, is_rag=is_rag_file, file_hash=file_hash)

        return {
            "success": True,
            "filename": safe_filename,
            "message": "Fichier �ph�m�re extrait et charg� dans la session avec succ�s."
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"�chec de l'extraction ou du stockage du fichier : {str(e)}"
        )


@app.get("/api/v1/usage/summary")
def get_tokens_usage_summary():
    """Consommation cumul�e de tokens."""
    try:
        from app.services.token_tracker import get_usage_summary
        return get_usage_summary()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"�chec de la r�cup�ration du r�sum� : {str(e)}")


@app.get("/api/v1/usage/history")
def get_tokens_usage_history(limit: int = 50):
    """Historique des appels LLM."""
    try:
        from app.services.token_tracker import get_usage_history
        return get_usage_history(limit)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Échec de la récupération de l'historique : {str(e)}")


@app.delete("/api/v1/admin/purge-chunks")
def purge_chunks_admin(
    session_id: Optional[str] = None,
    file_hash: Optional[str] = None,
    older_than_days: Optional[int] = None
):
    """
    Endpoint d'administration pour nettoyer les chunks obsolètes dans pgvector.
    Peut purger par session_id, par file_hash ou par ancienneté (older_than_days).
    """
    from app.rag.retriever import (
        purge_chunks_by_session_id,
        purge_chunks_by_file_hash,
        purge_old_session_chunks
    )
    deleted_total = 0
    if session_id:
        deleted_total += purge_chunks_by_session_id(session_id)
    if file_hash:
        deleted_total += purge_chunks_by_file_hash(file_hash)
    if older_than_days is not None:
        deleted_total += purge_old_session_chunks(older_than_days)

    if not session_id and not file_hash and older_than_days is None:
        # Défaut : purge des sessions inutilisées depuis plus de 7 jours
        deleted_total += purge_old_session_chunks(days=7)

    return {
        "success": True,
        "deleted_chunks_count": deleted_total,
        "message": f"{deleted_total} chunks ont été supprimés de la base vectorielle."
    }


@app.get("/api/v1/specs/library")
def get_specs_library():
    """
    Retourne la liste des documents de spécification déjà indexés en base vectorielle,
    dédoublonnnés par file_hash, avec nom de fichier d'origine et date d'indexation.
    """
    try:
        from app.rag.retriever import list_spec_library
        docs = list_spec_library()
        return {"success": True, "documents": docs, "count": len(docs)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Impossible de charger la bibliothèque de spécifications : {str(e)}")


# =========================================================================
# PARTIE A : ROUTES HISTORIQUE DE CONVERSATIONS
# =========================================================================

class ConversationCreateRequest(BaseModel):
    agent_type: str = "logs"  # "logs" ou "docs"
    title: str = "Nouvelle conversation"
    messages: Optional[List[dict]] = Field(default_factory=list)
    result: Optional[dict] = None

class ConversationUpdateRequest(BaseModel):
    title: Optional[str] = None
    messages: Optional[List[dict]] = None
    result: Optional[dict] = None


@app.get("/api/v1/conversations")
def list_conversations(agent: str = "logs", limit: int = 10):
    """Retourne les {limit} conversations les plus récentes pour un agent (logs ou docs)."""
    from app.services.conversation_history import get_conversations
    convs = get_conversations(agent_type=agent, limit=limit)
    return {"conversations": convs, "count": len(convs)}


@app.get("/api/v1/conversations/{conv_id}")
def get_conversation_detail(conv_id: str):
    """Retourne le détail complet d'une conversation (messages + résultat)."""
    from app.services.conversation_history import get_conversation
    conv = get_conversation(conv_id)
    if not conv:
        raise HTTPException(status_code=404, detail="Conversation non trouvée.")
    return conv


@app.post("/api/v1/conversations")
def create_conversation_endpoint(req: ConversationCreateRequest):
    """Crée une nouvelle conversation."""
    from app.services.conversation_history import create_conversation
    conv = create_conversation(
        agent_type=req.agent_type,
        title=req.title,
        messages=req.messages,
        result=req.result,
    )
    return conv


@app.put("/api/v1/conversations/{conv_id}")
def update_conversation_endpoint(conv_id: str, req: ConversationUpdateRequest):
    """Met à jour une conversation existante (titre, messages, résultat)."""
    from app.services.conversation_history import update_conversation
    conv = update_conversation(
        conv_id=conv_id,
        title=req.title,
        messages=req.messages,
        result=req.result,
    )
    if not conv:
        raise HTTPException(status_code=404, detail="Conversation non trouvée.")
    return conv


@app.delete("/api/v1/conversations/{conv_id}")
def delete_conversation_endpoint(conv_id: str):
    """Supprime une conversation."""
    from app.services.conversation_history import delete_conversation
    deleted = delete_conversation(conv_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="Conversation non trouvée.")
    return {"success": True, "deleted_id": conv_id}


@app.get("/api/v1/functions/{function_name}/doc")
async def get_function_doc(function_name: str, session_id: Optional[str] = None):
    """
    Retourne la documentation d'une fonction PowerCARD enrichie par LLM (4 sections structurées).
    Utilise un cache mémoire par (function_name, session_id) pour éviter les appels LLM répétés.

    Structure de réponse :
    - found: bool
    - llm_structured: { description, call_context, failure_meaning, diagnostic_hint }
    - raw_sources_count: int
    - message: str (si not found)
    """
    if not function_name or not function_name.strip():
        raise HTTPException(status_code=400, detail="Le nom de la fonction est requis.")

    safe_name = function_name.strip()
    cache_key = (safe_name, (session_id or "").strip())

    # Vérifier le cache mémoire
    if cache_key in _FUNC_DOC_CACHE:
        return _FUNC_DOC_CACHE[cache_key]

    raw_content_parts = []

    # 1. Interrogation du document de session (RAG vectoriel) si session_id fourni
    if session_id and session_id.strip():
        try:
            from app.rag.retriever import get_session_vectorstore
            session_db = get_session_vectorstore()
            query = f"{safe_name} fonction erreur HSM PowerCARD"
            session_docs = await run_in_threadpool(
                session_db.similarity_search, query, 3,
                {"session_id": session_id.strip()}
            )
            if session_docs:
                snippets = []
                for d in session_docs:
                    src = d.metadata.get("source", "Document de session")
                    pg = d.metadata.get("page")
                    pg_str = f" (p. {pg})" if pg else ""
                    snippets.append(f"[{src}{pg_str}]\n{d.page_content.strip()[:600]}")
                raw_content_parts.append("\n\n".join(snippets))
        except Exception as e:
            print(f"[WARN] get_function_doc session RAG failed for '{safe_name}': {e}")

    # 2. Fallback : Spec_PowerCARD.xlsx
    try:
        from app.rag.retriever import _local_excel_fallback
        excel_result = await run_in_threadpool(_local_excel_fallback, safe_name, 3)
        if excel_result and excel_result.strip():
            raw_content_parts.append(f"[Spec_PowerCARD.xlsx]\n{excel_result.strip()}")
    except Exception as e:
        print(f"[WARN] get_function_doc Excel fallback failed for '{safe_name}': {e}")

    # Cas : aucune documentation trouvée
    if not raw_content_parts:
        result = {
            "function_name": safe_name,
            "found": False,
            "llm_structured": None,
            "raw_sources_count": 0,
            "message": f"Documentation non disponible pour la fonction '{safe_name}' — elle n'est pas répertoriée dans Spec_PowerCARD.xlsx ni dans le document de session."
        }
        _FUNC_DOC_CACHE[cache_key] = result
        return result

    # 3. Enrichissement LLM : transformer la doc brute en 4 sections structurées
    combined_raw = "\n\n---\n\n".join(raw_content_parts)
    llm_structured = None

    try:
        from app.core.agent_graph import llm
        from app.services.llm_util import invoke_llm_with_retry
        from langchain_core.prompts import ChatPromptTemplate
        import json as _json

        doc_prompt = ChatPromptTemplate.from_messages([
            ("system", (
                "Tu es un expert en analyse de spécifications de systèmes monétiques PowerCARD (HPS).\n"
                "À partir de la documentation brute fournie sur une fonction PowerCARD, génère un OBJET JSON STRICTEMENT VALIDE "
                "avec exactement ces 4 clés (rien d'autre, pas de markdown, pas de texte en dehors du JSON) :\n"
                # CORRECTIF : les accolades JSON littérales doivent être doublées ({{ }}) dans un
                # template LangChain pour ne pas être interprétées comme des variables de template.
                "{{\n"
                '  "description": "<1-2 phrases : que fait cette fonction dans le système PowerCARD>",\n'
                '  "call_context": "<1-2 phrases : dans quel flux ou contexte cette fonction est-elle appelée>",\n'
                '  "failure_meaning": "<1-2 phrases : que signifie un résultat -1, -2 ou NOK pour cette fonction précise>",\n'
                '  "diagnostic_hint": "<1-2 phrases : quelle est la première chose à vérifier en cas d\'échec de cette fonction>"\n'
                "}}\n"
                "Si la documentation ne contient pas suffisamment d'information pour remplir une section, "
                "indique 'Information non disponible dans la documentation fournie.' pour cette section.\n"
                "RÉPONDS UNIQUEMENT avec l'objet JSON valide, sans aucun texte autour."
            )),
            ("user", (
                # CORRECTIF : ne jamais mélanger f-string Python et variables LangChain {var} dans
                # la même chaîne. Passer function_name comme variable nommée via format_messages().
                "Fonction : {function_name}\n\n"
                "Documentation brute :\n{raw_doc}"
            )),
        ])

        llm_resp = invoke_llm_with_retry(
            llm,
            doc_prompt.format_messages(function_name=safe_name, raw_doc=combined_raw[:3000])
        )
        resp_text = str(getattr(llm_resp, "content", llm_resp) or "").strip()
        if resp_text.startswith("```"):
            resp_text = re.sub(r"^```(?:json)?\s*", "", resp_text, flags=re.IGNORECASE)
            resp_text = re.sub(r"\s*```$", "", resp_text).strip()
        parsed = _json.loads(resp_text)
        if isinstance(parsed, dict) and "description" in parsed:
            llm_structured = parsed
    except Exception as e:
        print(f"[WARN] get_function_doc LLM enrichment failed for '{safe_name}': {e}")

    result = {
        "function_name": safe_name,
        "found": True,
        "llm_structured": llm_structured,
        "documentation": [
            {
                "source": part.split("\n")[0].strip("[]") if part.startswith("[") else "Source Spécification",
                "content": "\n".join(part.split("\n")[1:]).strip() if part.startswith("[") else part
            }
            for part in raw_content_parts
        ],
        "raw_sources_count": len(raw_content_parts),
    }

    # Mettre en cache uniquement si on a trouvé des infos
    _FUNC_DOC_CACHE[cache_key] = result
    return result