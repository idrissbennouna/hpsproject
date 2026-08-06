# backend/app/rag/loaders.py
import os
from pathlib import Path
from langchain_core.documents import Document
import openpyxl

def load_excel_specs(excel_path: Path) -> list[Document]:
    """
    Convertit chaque fonction documentée de Spec_PowerCARD.xlsx en Document réel.
    Cette fonction est extraite de ingest_docs.py pour être partagée.
    """
    excel_path = Path(excel_path)
    if not excel_path.exists():
        print(f"⚠️ Fichier Excel introuvable : {excel_path} — ignoré.")
        return []

    wb = openpyxl.load_workbook(str(excel_path), data_only=True)
    if "Lib" in wb.sheetnames:
        ws = wb["Lib"]
    else:
        ws = wb[wb.sheetnames[0]]
        print(f"⚠️ Onglet 'Lib' introuvable dans {excel_path.name}, utilisation du premier onglet '{ws.title}' à la place.")

    documents = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        func_name, source, path, description, exception = row[:5]
        if func_name:
            content = (
                f"Fonction {func_name.strip()} (module {source}, fichier {path})\n"
                f"Description : {description}\n"
                f"Conditions : {exception}"
            )
            documents.append(Document(
                page_content=content,
                metadata={
                    "source_file": "Spec_PowerCARD.xlsx", 
                    "function": func_name.strip(),
                    "module": str(source).strip() if source else "",
                    "file_path": str(path).strip() if path else ""
                }
            ))
    print(f"📊 {len(documents)} fonctions extraites de Spec_PowerCARD.xlsx.")
    return documents

def load_pdf_field_sections(pdf_path: Path) -> list[Document]:
    """
    Découpe le PDF de spécifications par section de champ numérotée de type 'Field N—...'.
    Extrait les documents et conserve les métadonnées de page et de champ.
    """
    import re
    pdf_path = Path(pdf_path)
    if not pdf_path.exists():
        print(f"⚠️ Fichier PDF introuvable : {pdf_path} — ignoré.")
        return []

    try:
        import pdfplumber
    except ImportError:
        print("⚠️ pdfplumber est requis pour parser les sections du PDF.")
        return []

    print(f"📄 Analyse et découpage par sections du PDF : {pdf_path.name}...")
    documents = []
    
    current_field_num = None
    current_field_name = None
    current_content = []
    current_page = None

    try:
        with pdfplumber.open(str(pdf_path)) as pdf:
            for page_num, page in enumerate(pdf.pages, 1):
                text = page.extract_text() or ""
                for line in text.splitlines():
                    # Détecte le début d'une section de champ
                    match = re.search(r"Field\s+(\d{1,3})\s*[—\-–]+\s*(.+)", line, re.IGNORECASE)
                    if match:
                        if current_field_num and current_content:
                            doc_content = "\n".join(current_content)
                            documents.append(Document(
                                page_content=doc_content,
                                metadata={
                                    "source": pdf_path.name,
                                    "source_file": pdf_path.name,
                                    "field_number": current_field_num,
                                    "field_name": current_field_name,
                                    "page": current_page,
                                    "attributes": ""  # Sera extrait si disponible lors de la validation
                                }
                            ))
                        current_field_num = match.group(1)
                        current_field_name = match.group(2).strip()
                        current_page = page_num
                        current_content = [line]
                    else:
                        if current_field_num:
                            current_content.append(line)

            # Ajouter la dernière section
            if current_field_num and current_content:
                doc_content = "\n".join(current_content)
                documents.append(Document(
                    page_content=doc_content,
                    metadata={
                        "source": pdf_path.name,
                        "source_file": pdf_path.name,
                        "field_number": current_field_num,
                        "field_name": current_field_name,
                        "page": current_page,
                        "attributes": ""
                    }
                ))
        print(f"📊 {len(documents)} sections de champs extraites du PDF {pdf_path.name}.")
    except Exception as e:
        print(f"❌ Erreur lors du découpage du PDF {pdf_path.name} : {e}")

    return documents

def load_all_docs(storage_dir: Path) -> list[Document]:
    """
    Parcourt le dossier de stockage et charge tous les documents supportés.
    """
    storage_dir = Path(storage_dir)
    documents = []
    
    # Ingestion Excel
    excel_path = storage_dir / "Spec_PowerCARD.xlsx"
    if excel_path.exists():
        documents.extend(load_excel_specs(excel_path))
    else:
        print(f"⚠️ Fichier Spec_PowerCARD.xlsx introuvable dans {storage_dir}")
        
    # Ingestion PDF (Stub)
    for file_path in storage_dir.glob("*.pdf"):
        documents.extend(load_pdf_field_sections(file_path))
        
    return documents
