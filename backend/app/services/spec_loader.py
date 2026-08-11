"""
Chargement DYNAMIQUE de la spécification métier PowerCARD (Spec_PowerCARD.xlsx).

Ce module remplace toute liste ou tout texte "en dur" concernant les fonctions
métier (CardInSaf, GetOriginalAuthData, CheckLimits, ...) par une lecture directe
du fichier Excel de spécification fourni par l'équipe HPS.

Tant que la base pgvector (RAG de YZ) n'est pas branchée, ce module sert de
"source de vérité" locale pour :
  - la liste des fonctions métier à surveiller dans les traces (plus besoin de
    les recopier à la main dans log_parser.py et agent_graph.py)
  - le texte de spécification (description + codes retour) à injecter dans le
    contexte du LLM, à la place des chaînes de texte écrites manuellement
"""

import logging
import os
import functools
import openpyxl
from pathlib import Path

logger = logging.getLogger(__name__)

_BACKEND_DIR = Path(__file__).resolve().parents[2]
DEFAULT_SPEC_PATH = os.getenv(
    "POWERCARD_SPEC_PATH",
    str(_BACKEND_DIR / "app" / "storage" / "Spec_PowerCARD.xlsx"),
)


@functools.lru_cache(maxsize=4)
def load_function_specs(spec_path: str = None) -> dict:
    """
    Lit la feuille 'Lib' du fichier Excel de spécification et retourne :
        { "NomFonction": {"source", "path", "description", "exception"} }

    Résultat mis en cache (le fichier de spec ne change pas en cours d'exécution).
    Utiliser load_function_specs.cache_clear() si le fichier est modifié à chaud.

    ROBUSTESSE : Si le fichier est absent, corrompu, ou si la feuille 'Lib' manque,
    un WARNING est loggué et un dict vide est retourné (et mis en cache).
    Le serveur ne plante JAMAIS à cause de ce fichier.
    """
    path = spec_path or DEFAULT_SPEC_PATH
    if not os.path.exists(path):
        logger.warning(
            "[spec_loader] Fichier de spécification introuvable : '%s'. "
            "Le pipeline continuera sans données Excel — seuls RAG session et LLM seront utilisés. "
            "Définir POWERCARD_SPEC_PATH si le fichier est dans un dossier non-standard.",
            path,
        )
        return {}

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        logger.warning(
            "[spec_loader] Impossible d'ouvrir '%s' (fichier corrompu ou format inconnu) : %s. "
            "Le pipeline continuera sans données Excel.",
            path, e,
        )
        return {}

    try:
        ws = wb["Lib"]
    except KeyError:
        logger.warning(
            "[spec_loader] La feuille 'Lib' est absente dans '%s'. "
            "Feuilles disponibles : %s. Le pipeline continuera sans données Excel.",
            path, wb.sheetnames,
        )
        return {}
    except Exception as e:
        logger.warning(
            "[spec_loader] Erreur lors de l'accès à la feuille 'Lib' dans '%s' : %s.",
            path, e,
        )
        return {}

    specs = {}
    try:
        for row in ws.iter_rows(min_row=2, values_only=True):
            padded = list(row) + [None] * 5
            func_name, source, rel_path, description, exception = padded[:5]

            if not func_name:
                continue

            clean_name = str(func_name).strip()
            specs[clean_name] = {
                "source": str(source).strip() if source else "",
                "path": str(rel_path).strip() if rel_path else "",
                "description": str(description).strip() if description else "",
                "exception": str(exception).strip() if exception else "",
            }
    except Exception as e:
        logger.warning(
            "[spec_loader] Erreur lors de la lecture des lignes de '%s' : %s. "
            "Données partiellement chargées (%d fonctions).",
            path, e, len(specs),
        )

    logger.info("[spec_loader] %d fonctions chargées depuis '%s'.", len(specs), path)
    return specs


def get_monitored_function_names(spec_path: str = None) -> list:
    """Liste des fonctions métier à surveiller, dérivée dynamiquement du fichier de spec."""
    return list(load_function_specs(spec_path).keys())


def get_spec_context_for_functions(func_names, spec_path: str = None) -> str:
    """
    Construit un texte de contexte à partir des vraies entrées de spec pour
    les fonctions données. Sert de RAG local en attendant la vraie base
    pgvector de YZ.
    """
    specs = load_function_specs(spec_path)
    blocks = []
    for name in func_names:
        entry = specs.get(name)
        if not entry:
            continue
        blocks.append(
            f"### Fonction {name} (source: {entry['source']}, path: {entry['path']})\n"
            f"Description : {entry['description']}\n"
            f"Codes retour / Exceptions :\n{entry['exception']}"
        )
    return "\n\n".join(blocks)